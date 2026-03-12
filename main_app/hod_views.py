import json
from itertools import groupby

import openpyxl
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt

from .forms import (
    AdminForm,
    CourseForm,
    ParentEditForm,
    ParentForm,
    SessionForm,
    StaffForm,
    StudentForm,
    SubjectForm,
)
from .models import (
    Admin,
    Attendance,
    AttendanceReport,
    Course,
    CustomUser,
    FeedbackStaff,
    FeedbackStudent,
    LeaveReportStaff,
    LeaveReportStudent,
    NotificationParent,
    NotificationStaff,
    NotificationStudent,
    Parent,
    ParentFeedback,
    Session,
    Staff,
    Student,
    StudentResult,
    Subject,
)


def _save_profile_picture(uploaded_file):
    if not uploaded_file:
        return None
    fs = FileSystemStorage()
    filename = fs.save(uploaded_file.name, uploaded_file)
    return fs.url(filename)


def _notify_parents_for_student(student, message):
    parents = Parent.objects.filter(student=student)
    notifications = [
        NotificationParent(parent=parent, message=message)
        for parent in parents
    ]
    if notifications:
        NotificationParent.objects.bulk_create(notifications)


def _delete_staff_academic_dependencies(staff_queryset):
    staff_ids = list(staff_queryset.values_list('id', flat=True))
    if not staff_ids:
        return

    subject_ids = list(Subject.objects.filter(staff_id__in=staff_ids).values_list('id', flat=True))
    if subject_ids:
        attendance_ids = list(Attendance.objects.filter(subject_id__in=subject_ids).values_list('id', flat=True))
        if attendance_ids:
            AttendanceReport.objects.filter(attendance_id__in=attendance_ids).delete()
            Attendance.objects.filter(id__in=attendance_ids).delete()
        StudentResult.objects.filter(subject_id__in=subject_ids).delete()
        Subject.objects.filter(id__in=subject_ids).delete()


def _resolve_session_value(session_value):
    if session_value in (None, ''):
        return None

    if isinstance(session_value, Session):
        return session_value

    session_text = str(session_value).strip()
    if not session_text:
        return None

    if session_text.isdigit():
        return Session.objects.filter(id=int(session_text)).first()

    session = Session.objects.filter(name__iexact=session_text).first()
    if session:
        return session

    session = Session.objects.filter(start_year=session_text).first()
    if session:
        return session

    if ' to ' in session_text:
        normalized = session_text
        if normalized.startswith('From '):
            normalized = normalized[5:]
        start_text, end_text = [part.strip() for part in normalized.split(' to ', 1)]
        return Session.objects.filter(start_year=start_text, end_year=end_text).first()

    return Session.objects.filter(start_year=session_text).first()


def _resolve_subject_value(subject_value):
    if subject_value in (None, ''):
        return None

    if isinstance(subject_value, Subject):
        return subject_value

    subject_text = str(subject_value).strip()
    if not subject_text:
        return None

    if subject_text.isdigit():
        return Subject.objects.filter(id=int(subject_text)).first()

    return Subject.objects.filter(name__iexact=subject_text).first()


def admin_home(request):
    total_staff = Staff.objects.count()
    total_students = Student.objects.count()
    subjects = Subject.objects.all()
    total_subject = subjects.count()
    total_course = Course.objects.count()
    attendance_totals = [Attendance.objects.filter(subject=subject).count() for subject in subjects]

    course_name_list = []
    student_count_list_in_course = []
    for course in Course.objects.all():
        course_name_list.append(course.name)
        student_count_list_in_course.append(Student.objects.filter(course=course).count())

    subject_list = []
    student_count_list_in_subject = []
    for subject in subjects:
        subject_list.append(subject.name)
        student_count_list_in_subject.append(Student.objects.filter(course=subject.course).count())

    student_attendance_present_list = []
    student_attendance_leave_list = []
    student_name_list = []
    for student in Student.objects.select_related('admin').all():
        present = AttendanceReport.objects.filter(student=student, status=True).count()
        absent = AttendanceReport.objects.filter(student=student, status=False).count()
        approved_leave = LeaveReportStudent.objects.filter(student=student, status=1).count()
        student_attendance_present_list.append(present)
        student_attendance_leave_list.append(absent + approved_leave)
        student_name_list.append(student.admin.first_name)

    context = {
        'page_title': "Administrative Dashboard",
        'total_students': total_students,
        'total_staff': total_staff,
        'total_course': total_course,
        'total_subject': total_subject,
        'subject_list': subject_list,
        'attendance_list': attendance_totals,
        'student_attendance_present_list': student_attendance_present_list,
        'student_attendance_leave_list': student_attendance_leave_list,
        'student_name_list': student_name_list,
        'student_count_list_in_subject': student_count_list_in_subject,
        'student_count_list_in_course': student_count_list_in_course,
        'course_name_list': course_name_list,
    }
    return render(request, 'hod_template/home_content.html', context)


def add_staff(request):
    form = StaffForm(request.POST or None, request.FILES or None)
    context = {
        'form': form,
        'page_title': 'Add Staff',
        'sample_file_url': '/media/samples/staff_sample.xlsx',
        'sample_file_name': 'staff_sample.xlsx',
        'upload_url_name': 'add_staff_from_xlsx',
        'upload_title': 'Bulk Add Staff',
        'upload_hint': 'Columns: first_name, last_name, address, email, gender, password, course_name',
        'upload_note': 'Use an existing course name exactly as it appears in Manage Courses.',
    }
    if request.method == 'POST':
        if form.is_valid():
            try:
                profile_pic = _save_profile_picture(request.FILES.get('profile_pic'))
                user = CustomUser.objects.create_user(
                    email=form.cleaned_data['email'],
                    password=form.cleaned_data['password'],
                    user_type='2',
                    first_name=form.cleaned_data['first_name'],
                    last_name=form.cleaned_data['last_name'],
                    profile_pic=profile_pic or '',
                )
                user.gender = form.cleaned_data['gender']
                user.address = form.cleaned_data['address']
                user.staff.course = form.cleaned_data['course']
                user.save()
                messages.success(request, "Successfully Added")
                return redirect(reverse('add_staff'))
            except Exception as e:
                messages.error(request, f"Could not add staff: {e}")
        else:
            messages.error(request, "Please fulfil all requirements")

    return render(request, 'hod_template/add_staff_template.html', context)


def add_staff_from_xlsx(request):
    if request.method == "POST" and request.FILES.get('xlsx_file'):
        workbook = openpyxl.load_workbook(request.FILES['xlsx_file'])
        sheet = workbook.active
        created_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 7:
                messages.error(request, f"Incomplete data in row: {row}")
                continue

            first_name, last_name, address, email, gender, password, course_name = row[:7]
            course = Course.objects.filter(name=course_name).first()
            if course is None:
                messages.error(request, f"Course '{course_name}' not found.")
                continue
            if CustomUser.objects.filter(email=email).exists():
                messages.error(request, f"Email {email} already exists.")
                continue

            try:
                user = CustomUser.objects.create_user(
                    email=email,
                    password=password,
                    user_type='2',
                    first_name=first_name,
                    last_name=last_name,
                    profile_pic='',
                )
                user.gender = gender
                user.address = address
                user.staff.course = course
                user.save()
                created_count += 1
            except Exception as e:
                messages.error(request, f"Error creating user {email}: {e}")

        messages.success(request, f"{created_count} staff accounts created successfully from XLSX.")
        return redirect('add_staff')

    return render(
        request,
        'hod_template/add_staff_template.html',
        {
            'form': StaffForm(),
            'page_title': 'Add Staff',
            'sample_file_url': '/media/samples/staff_sample.xlsx',
            'sample_file_name': 'staff_sample.xlsx',
            'upload_url_name': 'add_staff_from_xlsx',
            'upload_title': 'Bulk Add Staff',
            'upload_hint': 'Columns: first_name, last_name, address, email, gender, password, course_name',
            'upload_note': 'Use an existing course name exactly as it appears in Manage Courses.',
        },
    )


def add_parent(request):
    form = ParentForm(request.POST or None, request.FILES or None)
    context = {'form': form, 'page_title': 'Add Parent'}

    if request.method == 'POST':
        if form.is_valid():
            try:
                linked_student = Student.objects.get(roll_number=form.cleaned_data['student_roll_number'])
                profile_pic = _save_profile_picture(request.FILES.get('profile_pic'))
                user = CustomUser.objects.create_user(
                    email=form.cleaned_data['email'],
                    password=form.cleaned_data['password'],
                    user_type='4',
                    first_name=form.cleaned_data['first_name'],
                    last_name=form.cleaned_data['last_name'],
                    profile_pic=profile_pic or '',
                )
                user.gender = form.cleaned_data['gender']
                user.address = form.cleaned_data['address']
                user.parent.student = linked_student
                user.save()
                messages.success(request, "Parent Successfully Added")
                return redirect(reverse('add_parent'))
            except Exception as e:
                messages.error(request, f"Could not add parent: {e}")
        else:
            messages.error(request, "Please fulfil all requirements")

    context.update(
        {
            'sample_file_url': '/media/samples/parents_sample.xlsx',
            'sample_file_name': 'parents_sample.xlsx',
            'upload_url_name': 'add_parents_from_xlsx',
            'upload_title': 'Bulk Add Parents',
            'upload_hint': 'Columns: first_name, last_name, address, email, gender, password, student_roll_number',
            'upload_note': 'Import students first, then parents using matching student_roll_number values from the student sample.',
        }
    )
    return render(request, 'hod_template/add_parent_template.html', context)


def add_student(request):
    student_form = StudentForm(request.POST or None, request.FILES or None)
    context = {'form': student_form, 'page_title': 'Add Student'}
    if request.method == 'POST':
        if student_form.is_valid():
            try:
                profile_pic = _save_profile_picture(request.FILES.get('profile_pic'))
                user = CustomUser.objects.create_user(
                    email=student_form.cleaned_data['email'],
                    password=student_form.cleaned_data['password'],
                    user_type='3',
                    first_name=student_form.cleaned_data['first_name'],
                    last_name=student_form.cleaned_data['last_name'],
                    profile_pic=profile_pic or '',
                )
                user.gender = student_form.cleaned_data['gender']
                user.address = student_form.cleaned_data['address']
                user.student.session = student_form.cleaned_data['session']
                user.student.course = student_form.cleaned_data['course']
                user.student.roll_number = student_form.cleaned_data['roll_number']
                user.save()
                messages.success(request, "Successfully Added")
                return redirect(reverse('add_student'))
            except Exception as e:
                messages.error(request, f"Could Not Add: {e}")
        else:
            messages.error(request, "Could not add student")
    context.update(
        {
            'sample_file_url': '/media/samples/students_sample.xlsx',
            'sample_file_name': 'students_sample.xlsx',
            'upload_url_name': 'add_students_from_xlsx',
            'upload_title': 'Bulk Add Students',
            'upload_hint': 'Columns: first_name, last_name, address, email, gender, password, roll_number, course_name, section_name(or section_id)',
            'upload_note': 'This sample is paired with parents_sample.xlsx. Import students before parents and use the section name exactly as it appears in Manage Sections.',
        }
    )
    return render(request, 'hod_template/add_student_template.html', context)


def add_students_from_xlsx(request):
    if request.method == 'POST' and request.FILES.get('xlsx_file'):
        workbook = openpyxl.load_workbook(request.FILES['xlsx_file'])
        sheet = workbook.active
        created_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 8:
                messages.error(
                    request,
                    "Student upload requires at least: first_name, last_name, address, email, gender, password, roll_number, course_name",
                )
                continue

            first_name, last_name, address, email, gender, password, roll_number, course_name, *rest = row
            session_value = rest[0] if rest else None

            course = Course.objects.filter(name=course_name).first()
            session = _resolve_session_value(session_value)
            if course is None:
                messages.error(request, f"Course '{course_name}' not found.")
                continue
            if session_value not in (None, '') and session is None:
                messages.error(
                    request,
                    f"Section '{session_value}' not found. Use a valid section id or section name.",
                )
                continue
            if CustomUser.objects.filter(email=email).exists():
                messages.error(request, f"Email {email} already exists.")
                continue
            if Student.objects.filter(roll_number=roll_number).exists():
                messages.error(request, f"Roll number {roll_number} already exists.")
                continue

            try:
                user = CustomUser.objects.create_user(
                    email=email,
                    password=password,
                    user_type='3',
                    first_name=first_name,
                    last_name=last_name,
                    profile_pic='',
                )
                user.gender = gender
                user.address = address
                user.student.course = course
                user.student.session = session
                user.student.roll_number = roll_number
                user.save()
                created_count += 1
            except Exception as e:
                messages.error(request, f"Error adding student {first_name} {last_name}: {e}")

        messages.success(request, f"{created_count} student accounts created successfully from XLSX.")
        return redirect('add_student')

    return render(
        request,
        'hod_template/add_student_template.html',
        {
            'form': StudentForm(),
            'page_title': 'Add Student',
            'sample_file_url': '/media/samples/students_sample.xlsx',
            'sample_file_name': 'students_sample.xlsx',
            'upload_url_name': 'add_students_from_xlsx',
            'upload_title': 'Bulk Add Students',
            'upload_hint': 'Columns: first_name, last_name, address, email, gender, password, roll_number, course_name, section_name(or section_id)',
            'upload_note': 'This sample is paired with parents_sample.xlsx. Import students before parents and use the section name exactly as it appears in Manage Sections.',
        },
    )


def add_parents_from_xlsx(request):
    if request.method == 'POST' and request.FILES.get('xlsx_file'):
        workbook = openpyxl.load_workbook(request.FILES['xlsx_file'])
        sheet = workbook.active
        created_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 7:
                messages.error(
                    request,
                    "Parent upload requires: first_name, last_name, address, email, gender, password, student_roll_number",
                )
                continue

            first_name, last_name, address, email, gender, password, student_roll_number = row[:7]
            student = Student.objects.filter(roll_number=student_roll_number).first()
            if student is None:
                messages.error(request, f"Roll number '{student_roll_number}' not found.")
                continue
            if CustomUser.objects.filter(email=email).exists():
                messages.error(request, f"Email {email} already exists.")
                continue

            try:
                user = CustomUser.objects.create_user(
                    email=email,
                    password=password,
                    user_type='4',
                    first_name=first_name,
                    last_name=last_name,
                    profile_pic='',
                )
                user.gender = gender
                user.address = address
                user.parent.student = student
                user.save()
                created_count += 1
            except Exception as e:
                messages.error(request, f"Error creating parent {email}: {e}")

        messages.success(request, f"{created_count} parent accounts created successfully from XLSX.")
        return redirect('add_parent')

    return render(
        request,
        'hod_template/add_parent_template.html',
        {
            'form': ParentForm(),
            'page_title': 'Add Parent',
            'sample_file_url': '/media/samples/parents_sample.xlsx',
            'sample_file_name': 'parents_sample.xlsx',
            'upload_url_name': 'add_parents_from_xlsx',
            'upload_title': 'Bulk Add Parents',
            'upload_hint': 'Columns: first_name, last_name, address, email, gender, password, student_roll_number',
            'upload_note': 'Import students first, then parents using matching student_roll_number values from the student sample.',
        },
    )


def add_course(request):
    form = CourseForm(request.POST or None)
    context = {'form': form, 'page_title': 'Add Course'}
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Successfully Added")
            return redirect(reverse('add_course'))
        messages.error(request, "Could Not Add")
    return render(request, 'hod_template/add_course_template.html', context)


def add_subject(request):
    form = SubjectForm(request.POST or None)
    context = {'form': form, 'page_title': 'Add Subject'}
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Successfully Added")
            return redirect(reverse('add_subject'))
        messages.error(request, "Fill Form Properly")

    return render(request, 'hod_template/add_subject_template.html', context)


def manage_staff(request):
    context = {
        'allStaff': CustomUser.objects.filter(user_type='2'),
        'page_title': 'Manage Staff',
    }
    return render(request, "hod_template/manage_staff.html", context)


@csrf_exempt
def get_student_by_roll_number(request):
    roll_number = (request.POST.get('roll_number') or '').strip()
    student = Student.objects.filter(roll_number=roll_number).select_related('admin').first()
    if student is None:
        return JsonResponse({'found': False}, safe=True)
    return JsonResponse(
        {
            'found': True,
            'student_name': f"{student.admin.first_name} {student.admin.last_name}",
            'course': str(student.course) if student.course else '',
        }
    )


def manage_student(request):
    context = {
        'students': CustomUser.objects.filter(user_type='3').select_related('student__course', 'student__session'),
        'page_title': 'Manage Students',
    }
    return render(request, "hod_template/manage_student.html", context)


def manage_parent(request):
    context = {
        'parents': CustomUser.objects.filter(user_type='4').select_related('parent__student__admin'),
        'page_title': 'Manage Parents',
    }
    return render(request, "hod_template/manage_parent.html", context)


def manage_course(request):
    return render(
        request,
        "hod_template/manage_course.html",
        {'courses': Course.objects.all(), 'page_title': 'Manage Courses'},
    )


def manage_subject(request):
    return render(
        request,
        "hod_template/manage_subject.html",
        {'subjects': Subject.objects.all(), 'page_title': 'Manage Subjects'},
    )


def edit_staff(request, staff_id):
    staff = get_object_or_404(Staff, id=staff_id)
    form = StaffForm(request.POST or None, request.FILES or None, instance=staff)
    context = {'form': form, 'staff_id': staff_id, 'page_title': 'Edit Staff'}
    if request.method == 'POST':
        if form.is_valid():
            try:
                user = staff.admin
                profile_pic = _save_profile_picture(request.FILES.get('profile_pic'))
                user.email = form.cleaned_data['email']
                if form.cleaned_data.get('password'):
                    user.set_password(form.cleaned_data['password'])
                if profile_pic:
                    user.profile_pic = profile_pic
                user.first_name = form.cleaned_data['first_name']
                user.last_name = form.cleaned_data['last_name']
                user.gender = form.cleaned_data['gender']
                user.address = form.cleaned_data['address']
                staff.course = form.cleaned_data['course']
                user.save()
                staff.save()
                messages.success(request, "Successfully Updated")
                return redirect(reverse('edit_staff', args=[staff_id]))
            except Exception as e:
                messages.error(request, f"Could Not Update {e}")
        else:
            messages.error(request, "Please fill form properly")
    return render(request, "hod_template/edit_staff_template.html", context)


def edit_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    form = StudentForm(request.POST or None, request.FILES or None, instance=student)
    context = {'form': form, 'student_id': student_id, 'page_title': 'Edit Student'}
    if request.method == 'POST':
        if form.is_valid():
            try:
                user = student.admin
                profile_pic = _save_profile_picture(request.FILES.get('profile_pic'))
                user.email = form.cleaned_data['email']
                if form.cleaned_data.get('password'):
                    user.set_password(form.cleaned_data['password'])
                if profile_pic:
                    user.profile_pic = profile_pic
                user.first_name = form.cleaned_data['first_name']
                user.last_name = form.cleaned_data['last_name']
                user.gender = form.cleaned_data['gender']
                user.address = form.cleaned_data['address']
                student.session = form.cleaned_data['session']
                student.course = form.cleaned_data['course']
                student.roll_number = form.cleaned_data['roll_number']
                user.save()
                student.save()
                messages.success(request, "Successfully Updated")
                return redirect(reverse('edit_student', args=[student_id]))
            except Exception as e:
                messages.error(request, f"Could Not Update {e}")
        else:
            messages.error(request, "Please Fill Form Properly!")
    return render(request, "hod_template/edit_student_template.html", context)


def edit_parent(request, parent_id):
    parent = get_object_or_404(Parent, id=parent_id)
    form = ParentEditForm(request.POST or None, request.FILES or None, instance=parent)
    context = {'form': form, 'parent_id': parent_id, 'page_title': 'Edit Parent'}
    if request.method == 'POST':
        if form.is_valid():
            try:
                user = parent.admin
                profile_pic = _save_profile_picture(request.FILES.get('profile_pic'))
                linked_student = Student.objects.get(roll_number=form.cleaned_data['student_roll_number'])
                user.email = form.cleaned_data['email']
                if form.cleaned_data.get('password'):
                    user.set_password(form.cleaned_data['password'])
                if profile_pic:
                    user.profile_pic = profile_pic
                user.first_name = form.cleaned_data['first_name']
                user.last_name = form.cleaned_data['last_name']
                user.gender = form.cleaned_data['gender']
                user.address = form.cleaned_data['address']
                parent.student = linked_student
                user.save()
                parent.save()
                messages.success(request, "Parent updated successfully")
                return redirect(reverse('edit_parent', args=[parent_id]))
            except Exception as e:
                messages.error(request, f"Could not update parent: {e}")
        else:
            messages.error(request, "Please fill the form properly")
    return render(request, "hod_template/edit_parent_template.html", context)


def student_detail(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    attendance_reports = AttendanceReport.objects.filter(student=student).select_related('attendance__subject')
    results = StudentResult.objects.filter(student=student).exclude(
        result_type='semester'
    ).select_related('subject').order_by('result_type', 'assessment_name', 'subject__name')
    semester_results = StudentResult.objects.filter(
        student=student,
        result_type='semester',
    ).select_related('subject').order_by('assessment_name', 'subject__name')
    semester_groups = [
        {'semester_name': semester_name, 'results': list(group_results)}
        for semester_name, group_results in groupby(semester_results, key=lambda result: result.assessment_name)
    ]
    parents = Parent.objects.filter(student=student).select_related('admin')
    context = {
        'page_title': f"Student Details - {student.roll_number}",
        'student': student,
        'attendance_reports': attendance_reports,
        'results': results,
        'semester_groups': semester_groups,
        'parents': parents,
    }
    return render(request, "hod_template/student_detail.html", context)


def toggle_student_suspension(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    student.is_suspended = not student.is_suspended
    student.admin.is_active = not student.is_suspended
    student.admin.save()
    student.save()
    status_text = "suspended" if student.is_suspended else "reactivated"
    messages.success(request, f"Student {student.roll_number} has been {status_text}.")
    return redirect(reverse('manage_student'))


def edit_course(request, course_id):
    instance = get_object_or_404(Course, id=course_id)
    form = CourseForm(request.POST or None, instance=instance)
    context = {'form': form, 'course_id': course_id, 'page_title': 'Edit Course'}
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Successfully Updated")
        else:
            messages.error(request, "Could Not Update")

    return render(request, 'hod_template/edit_course_template.html', context)


def edit_subject(request, subject_id):
    instance = get_object_or_404(Subject, id=subject_id)
    form = SubjectForm(request.POST or None, instance=instance)
    context = {'form': form, 'subject_id': subject_id, 'page_title': 'Edit Subject'}
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Successfully Updated")
            return redirect(reverse('edit_subject', args=[subject_id]))
        messages.error(request, "Fill Form Properly")
    return render(request, 'hod_template/edit_subject_template.html', context)


def add_session(request):
    form = SessionForm(request.POST or None)
    context = {'form': form, 'page_title': 'Add Section'}
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Section Created")
            return redirect(reverse('add_session'))
        messages.error(request, 'Fill Form Properly')
    return render(request, "hod_template/add_session_template.html", context)


def manage_session(request):
    return render(
        request,
        "hod_template/manage_session.html",
        {'sessions': Session.objects.all(), 'page_title': 'Manage Sections'},
    )


def edit_session(request, session_id):
    instance = get_object_or_404(Session, id=session_id)
    form = SessionForm(request.POST or None, instance=instance)
    context = {'form': form, 'session_id': session_id, 'page_title': 'Edit Section'}
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Section Updated")
            return redirect(reverse('edit_session', args=[session_id]))
        messages.error(request, "Invalid Form Submitted")

    return render(request, "hod_template/edit_session_template.html", context)


@csrf_exempt
def check_email_availability(request):
    email = request.POST.get("email")
    return HttpResponse(CustomUser.objects.filter(email=email).exists())


@csrf_exempt
def student_feedback_message(request):
    if request.method != 'POST':
        return redirect(reverse('admin_feedbacks'))

    feedback = get_object_or_404(FeedbackStudent, id=request.POST.get('id'))
    feedback.reply = request.POST.get('reply')
    feedback.save()
    return HttpResponse(True)


@csrf_exempt
def parent_feedback_message(request):
    if request.method != 'POST':
        return redirect(reverse('admin_feedbacks'))

    feedback = get_object_or_404(ParentFeedback, id=request.POST.get('id'))
    feedback.reply = request.POST.get('reply')
    feedback.save()
    return HttpResponse("True")


@csrf_exempt
def staff_feedback_message(request):
    if request.method != 'POST':
        return redirect(reverse('admin_feedbacks'))

    feedback = get_object_or_404(FeedbackStaff, id=request.POST.get('id'))
    feedback.reply = request.POST.get('reply')
    feedback.save()
    return HttpResponse(True)


def admin_feedbacks(request):
    context = {
        'page_title': 'Feedback',
        'student_feedbacks': FeedbackStudent.objects.select_related('student__admin', 'student__session'),
        'staff_feedbacks': FeedbackStaff.objects.select_related('staff__admin', 'staff__course'),
        'parent_feedbacks': ParentFeedback.objects.select_related('parent__admin'),
    }
    return render(request, 'hod_template/feedbacks.html', context)


@csrf_exempt
def view_staff_leave(request):
    if request.method != 'POST':
        return redirect(reverse('admin_leaves'))

    leave = get_object_or_404(LeaveReportStaff, id=request.POST.get('id'))
    if leave.substitute_status != 1:
        return HttpResponse(False)
    leave.status = 1 if request.POST.get('status') == '1' else -1
    leave.save()
    NotificationStaff.objects.create(
        staff=leave.staff,
        message=(
            f"HOD {'approved' if leave.status == 1 else 'rejected'} your leave request for {leave.date}."
        ),
    )
    return HttpResponse(True)


@csrf_exempt
def view_student_leave(request):
    if request.method != 'POST':
        return redirect(reverse('admin_leaves'))

    leave = get_object_or_404(LeaveReportStudent, id=request.POST.get('id'))
    leave.status = 1 if request.POST.get('status') == '1' else -1
    leave.save()
    return HttpResponse(True)


def admin_leaves(request):
    context = {
        'page_title': 'Leave Requests',
        'staff_leaves': LeaveReportStaff.objects.select_related(
            'staff__admin',
            'staff__course',
            'substitute_staff__admin',
        ).filter(substitute_status=1),
        'student_leaves': LeaveReportStudent.objects.select_related('student__admin', 'student__course'),
    }
    return render(request, 'hod_template/leaves.html', context)


def admin_view_attendance(request):
    context = {
        'subjects': Subject.objects.all(),
        'sessions': Session.objects.all(),
        'page_title': 'View Attendance',
    }
    return render(request, "hod_template/admin_view_attendance.html", context)


@csrf_exempt
def get_admin_attendance(request):
    session = get_object_or_404(Session, id=request.POST.get('session'))
    attendance = get_object_or_404(Attendance, id=request.POST.get('attendance_date_id'), session=session)
    attendance_reports = AttendanceReport.objects.filter(attendance=attendance)
    json_data = [{"status": str(report.status), "name": str(report.student)} for report in attendance_reports]
    return JsonResponse(json_data, safe=False)


def admin_view_profile(request):
    admin = get_object_or_404(Admin, admin=request.user)
    form = AdminForm(request.POST or None, request.FILES or None, instance=admin)
    context = {'form': form, 'page_title': 'View/Edit Profile'}
    if request.method == 'POST':
        if form.is_valid():
            try:
                custom_user = admin.admin
                profile_pic = _save_profile_picture(request.FILES.get('profile_pic'))
                if form.cleaned_data.get('password'):
                    custom_user.set_password(form.cleaned_data['password'])
                if profile_pic:
                    custom_user.profile_pic = profile_pic
                custom_user.first_name = form.cleaned_data['first_name']
                custom_user.last_name = form.cleaned_data['last_name']
                custom_user.email = form.cleaned_data['email']
                custom_user.gender = form.cleaned_data['gender']
                custom_user.address = form.cleaned_data['address']
                custom_user.save()
                messages.success(request, "Profile Updated!")
                return redirect(reverse('admin_view_profile'))
            except Exception as e:
                messages.error(request, f"Error Occured While Updating Profile {e}")
        else:
            messages.error(request, "Invalid Data Provided")
    return render(request, "hod_template/admin_view_profile.html", context)


def admin_add_semester_result(request):
    context = {
        'page_title': "Publish Semester Results",
        'subjects': Subject.objects.all(),
        'sessions': Session.objects.all(),
        'sample_file_url': '/media/samples/semester_results_sample.xlsx',
        'sample_file_name': 'semester_results_sample.xlsx',
        'upload_url_name': 'add_semester_results_from_xlsx',
        'upload_title': 'Bulk Upload Semester Results',
        'upload_hint': 'Columns: roll_number, subject_name, section_name, semester_name, internal_score, external_score',
        'upload_note': 'Use the exact student roll number, subject name, and section name already present in the system.',
    }
    if request.method == 'POST':
        try:
            subject = get_object_or_404(Subject, id=request.POST.get('subject'))
            session = get_object_or_404(Session, id=request.POST.get('session'))
            assessment_name = (request.POST.get('assessment_name') or '').strip()
            if not assessment_name:
                raise ValueError("Semester name is required")
            students = Student.objects.filter(course=subject.course, session=session).select_related('admin')
            updated_count = 0

            for student in students:
                test_value = (request.POST.get(f'test_{student.id}') or '').strip()
                exam_value = (request.POST.get(f'exam_{student.id}') or '').strip()
                if test_value == '' and exam_value == '':
                    continue

                result, created = StudentResult.objects.update_or_create(
                    student=student,
                    subject=subject,
                    result_type='semester',
                    assessment_name=assessment_name,
                    defaults={
                        'test': float(test_value or 0),
                        'exam': float(exam_value or 0),
                    },
                )
                notification_text = (
                    f"Semester result {result.assessment_name} published for {student.roll_number} in {subject.name}. "
                    f"Internal: {result.test}, External: {result.exam}"
                )
                NotificationStudent.objects.create(student=student, message=notification_text)
                _notify_parents_for_student(student, notification_text)
                updated_count += 1

            if updated_count == 0:
                raise ValueError("Enter scores for at least one student")

            messages.success(request, f"Semester results saved for {updated_count} students.")
        except Exception as exc:
            messages.error(request, f"Could not publish semester result: {exc}")
    return render(request, "hod_template/admin_add_semester_result.html", context)


def add_semester_results_from_xlsx(request):
    context = {
        'page_title': "Publish Semester Results",
        'subjects': Subject.objects.all(),
        'sessions': Session.objects.all(),
        'sample_file_url': '/media/samples/semester_results_sample.xlsx',
        'sample_file_name': 'semester_results_sample.xlsx',
        'upload_url_name': 'add_semester_results_from_xlsx',
        'upload_title': 'Bulk Upload Semester Results',
        'upload_hint': 'Columns: roll_number, subject_name, section_name, semester_name, internal_score, external_score',
        'upload_note': 'Use the exact student roll number, subject name, and section name already present in the system.',
    }

    if request.method == 'POST' and request.FILES.get('xlsx_file'):
        workbook = openpyxl.load_workbook(request.FILES['xlsx_file'])
        sheet = workbook.active
        updated_count = 0

        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(value in (None, '') for value in row[:6]):
                continue

            if len(row) < 6:
                messages.error(
                    request,
                    f"Row {row_number} is incomplete. Required columns: roll_number, subject_name, section_name, semester_name, internal_score, external_score.",
                )
                continue

            roll_number, subject_value, session_value, assessment_name, test_value, exam_value = row[:6]
            student = Student.objects.filter(roll_number=str(roll_number).strip()).select_related('admin', 'session').first()
            subject = _resolve_subject_value(subject_value)
            session = _resolve_session_value(session_value)
            assessment_name = str(assessment_name or '').strip()

            if student is None:
                messages.error(request, f"Row {row_number}: roll number '{roll_number}' not found.")
                continue
            if subject is None:
                messages.error(request, f"Row {row_number}: subject '{subject_value}' not found.")
                continue
            if session is None:
                messages.error(request, f"Row {row_number}: section '{session_value}' not found.")
                continue
            if not assessment_name:
                messages.error(request, f"Row {row_number}: semester_name is required.")
                continue
            if student.course_id != subject.course_id:
                messages.error(
                    request,
                    f"Row {row_number}: subject '{subject.name}' does not belong to {student.roll_number}'s course.",
                )
                continue
            if student.session_id != session.id:
                messages.error(
                    request,
                    f"Row {row_number}: student {student.roll_number} is not assigned to section '{session.name}'.",
                )
                continue

            try:
                internal_score = float(test_value or 0)
                external_score = float(exam_value or 0)
            except (TypeError, ValueError):
                messages.error(request, f"Row {row_number}: internal_score and external_score must be numeric.")
                continue

            result, _ = StudentResult.objects.update_or_create(
                student=student,
                subject=subject,
                result_type='semester',
                assessment_name=assessment_name,
                defaults={
                    'test': internal_score,
                    'exam': external_score,
                },
            )
            notification_text = (
                f"Semester result {result.assessment_name} published for {student.roll_number} in {subject.name}. "
                f"Internal: {result.test}, External: {result.exam}"
            )
            NotificationStudent.objects.create(student=student, message=notification_text)
            _notify_parents_for_student(student, notification_text)
            updated_count += 1

        if updated_count:
            messages.success(request, f"Semester results imported for {updated_count} students.")
        else:
            messages.error(request, "No semester results were imported from the XLSX file.")

        return redirect('admin_add_semester_result')

    return render(request, "hod_template/admin_add_semester_result.html", context)


def admin_notifications(request):
    context = {
        'page_title': "Notifications",
        'students': CustomUser.objects.filter(user_type='3').select_related('student__course'),
        'allStaff': CustomUser.objects.filter(user_type='2').select_related('staff__course'),
        'parents': CustomUser.objects.filter(user_type='4').select_related('parent__student'),
    }
    return render(request, "hod_template/notifications.html", context)


@csrf_exempt
def send_parent_notification(request):
    if request.method != "POST":
        return JsonResponse("False", safe=False)

    try:
        parent = Parent.objects.get(admin__id=request.POST.get('id'))
        NotificationParent.objects.create(parent=parent, message=request.POST.get('message'))
        return JsonResponse("True", safe=False)
    except Exception:
        return JsonResponse("False", safe=False)


@csrf_exempt
def send_parent_notification_to_all(request):
    if request.method != "POST":
        return JsonResponse({"success": False}, safe=True)

    message = (request.POST.get('message') or '').strip()
    if not message:
        return JsonResponse({"success": False, "error": "Message is required."}, safe=True)

    notifications = [
        NotificationParent(parent=parent, message=message)
        for parent in Parent.objects.all()
    ]
    if notifications:
        NotificationParent.objects.bulk_create(notifications)
    return JsonResponse({"success": True, "count": len(notifications)}, safe=True)


@csrf_exempt
def send_notification_to_everyone(request):
    if request.method != "POST":
        return JsonResponse({"success": False}, safe=True)

    message = (request.POST.get('message') or '').strip()
    if not message:
        return JsonResponse({"success": False, "error": "Message is required."}, safe=True)

    staff_notifications = [
        NotificationStaff(staff=staff, message=message)
        for staff in Staff.objects.all()
    ]
    student_notifications = []
    parent_notifications = []
    students = list(Student.objects.select_related('admin'))

    for student in students:
        student_notifications.append(NotificationStudent(student=student, message=message))
        parents = Parent.objects.filter(student=student)
        parent_notifications.extend(
            NotificationParent(
                parent=parent,
                message=f"Student update for {student.roll_number}: {message}",
            )
            for parent in parents
        )

    parent_notifications.extend(
        NotificationParent(parent=parent, message=message)
        for parent in Parent.objects.filter(student__isnull=True)
    )

    if staff_notifications:
        NotificationStaff.objects.bulk_create(staff_notifications)
    if student_notifications:
        NotificationStudent.objects.bulk_create(student_notifications)
    if parent_notifications:
        NotificationParent.objects.bulk_create(parent_notifications)

    return JsonResponse(
        {
            "success": True,
            "staff_count": len(staff_notifications),
            "student_count": len(student_notifications),
            "parent_count": len(parent_notifications),
        },
        safe=True,
    )


@csrf_exempt
def send_student_notification(request):
    student = get_object_or_404(Student, admin_id=request.POST.get('id'))
    message = request.POST.get('message')
    NotificationStudent.objects.create(student=student, message=message)
    _notify_parents_for_student(student, f"Student update for {student.roll_number}: {message}")
    return HttpResponse("True")


@csrf_exempt
def send_student_notification_to_all(request):
    if request.method != "POST":
        return JsonResponse({"success": False}, safe=True)

    message = (request.POST.get('message') or '').strip()
    if not message:
        return JsonResponse({"success": False, "error": "Message is required."}, safe=True)

    students = list(Student.objects.select_related('admin'))
    notifications = [
        NotificationStudent(student=student, message=message)
        for student in students
    ]
    if notifications:
        NotificationStudent.objects.bulk_create(notifications)
    for student in students:
        _notify_parents_for_student(student, f"Student update for {student.roll_number}: {message}")
    return JsonResponse({"success": True, "count": len(notifications)}, safe=True)


@csrf_exempt
def send_staff_notification(request):
    staff = get_object_or_404(Staff, admin_id=request.POST.get('id'))
    NotificationStaff.objects.create(staff=staff, message=request.POST.get('message'))
    return HttpResponse("True")


@csrf_exempt
def send_staff_notification_to_all(request):
    if request.method != "POST":
        return JsonResponse({"success": False}, safe=True)

    message = (request.POST.get('message') or '').strip()
    if not message:
        return JsonResponse({"success": False, "error": "Message is required."}, safe=True)

    notifications = [
        NotificationStaff(staff=staff, message=message)
        for staff in Staff.objects.all()
    ]
    if notifications:
        NotificationStaff.objects.bulk_create(notifications)
    return JsonResponse({"success": True, "count": len(notifications)}, safe=True)


def delete_staff(request, staff_id):
    user = get_object_or_404(CustomUser, staff__id=staff_id)
    with transaction.atomic():
        _delete_staff_academic_dependencies(Staff.objects.filter(id=staff_id))
        user.delete()
    messages.success(request, "Staff deleted successfully!")
    return redirect(reverse('manage_staff'))


def delete_all_staff(request):
    if request.method == 'POST':
        staff_queryset = Staff.objects.all()
        with transaction.atomic():
            _delete_staff_academic_dependencies(staff_queryset)
            deleted_count, _ = CustomUser.objects.filter(user_type='2').delete()
        messages.success(request, f"Deleted all staff records. Removed {deleted_count} related entries.")
    return redirect(reverse('manage_staff'))


def delete_student(request, student_id):
    get_object_or_404(CustomUser, student__id=student_id).delete()
    messages.success(request, "Student deleted successfully!")
    return redirect(reverse('manage_student'))


def delete_all_students(request):
    if request.method == 'POST':
        deleted_count, _ = CustomUser.objects.filter(user_type='3').delete()
        messages.success(request, f"Deleted all student records. Removed {deleted_count} related entries.")
    return redirect(reverse('manage_student'))


def delete_parent(request, parent_id):
    get_object_or_404(CustomUser, parent__id=parent_id).delete()
    messages.success(request, "Parent deleted successfully!")
    return redirect(reverse('manage_parent'))


def delete_all_parents(request):
    if request.method == 'POST':
        deleted_count, _ = CustomUser.objects.filter(user_type='4').delete()
        messages.success(request, f"Deleted all parent records. Removed {deleted_count} related entries.")
    return redirect(reverse('manage_parent'))


def delete_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    try:
        course.delete()
        messages.success(request, "Course deleted successfully!")
    except Exception:
        messages.error(
            request,
            "Sorry, some students are assigned to this course already. Kindly change the affected student course and try again",
        )
    return redirect(reverse('manage_course'))


def delete_subject(request, subject_id):
    get_object_or_404(Subject, id=subject_id).delete()
    messages.success(request, "Subject deleted successfully!")
    return redirect(reverse('manage_subject'))


def delete_session(request, session_id):
    session = get_object_or_404(Session, id=session_id)
    try:
        session.delete()
        messages.success(request, "Section deleted successfully!")
    except Exception:
        messages.error(
            request,
            "There are students assigned to this section. Please move them to another section.",
        )
    return redirect(reverse('manage_session'))
